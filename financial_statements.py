import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList

def validate_data(data):
    """
    Validate the input data to ensure it contains all required keys and values.
    """
    required_keys = {
        'income': ['إجمالي الإيرادات | Total Revenue', 'صافي الربح | Net Profit'],
        'balance': ['إجمالي الأصول | Total Assets', 'إجمالي الخصوم | Total Liabilities', 'إجمالي حقوق الملكية | Total Equity'],
        'cash_flow': ['النقد وما في حكمه في نهاية السنة | Cash and cash equivalents at end of year']
    }
    errors = []
    for section, keys in required_keys.items():
        if section not in data:
            errors.append(f"المدخل '{section}' مفقود.")
            continue
        for key in keys:
            if key not in data[section]:
                errors.append(f"البند '{key}' مفقود في القسم '{section}'.")
            elif not isinstance(data[section][key].get('current', 0), (int, float)) or not isinstance(data[section][key].get('previous', 0), (int, float)):
                errors.append(f"القيم الخاصة بالبند '{key}' في القسم '{section}' ليست أرقامًا صالحة.")
    return errors

def generate_financial_statements(data, output_path):
    """Generate financial statements based on the provided data."""
    # Validate input data
    validation_errors = validate_data(data)
    if validation_errors:
        raise ValueError(f"أخطاء في البيانات المدخلة: {validation_errors}")
    
    wb = openpyxl.Workbook()
    # Create sheets for different financial statements
    sheets = {
        'تقرير عام | Overview': wb.active,
        'قائمة الدخل | Income Statement': wb.create_sheet(),
        'قائمة المركز المالي | Balance Sheet': wb.create_sheet(),
        'قائمة التغيرات في حقوق الملكية | Equity': wb.create_sheet(),
        'قائمة التدفقات النقدية | Cash Flow': wb.create_sheet(),
        'الملاحظات | Notes': wb.create_sheet(),
        'الرسوم البيانية | Charts': wb.create_sheet()
    }
    # Rename the default sheet
    sheets['تقرير عام | Overview'].title = 'تقرير عام | Overview'
    
    # Generate each statement
    generate_overview(sheets['تقرير عام | Overview'], data)
    generate_income_statement(sheets['قائمة الدخل | Income Statement'], data['income'])
    generate_balance_sheet(sheets['قائمة المركز المالي | Balance Sheet'], data['balance'])
    generate_equity_statement(sheets['قائمة التغيرات في حقوق الملكية | Equity'], data['equity'])
    generate_cash_flow_statement(sheets['قائمة التدفقات النقدية | Cash Flow'], data['cash_flow'])
    generate_notes(sheets['الملاحظات | Notes'], data.get('notes', {}))
    generate_charts(sheets['الرسوم البيانية | Charts'], data)
    
    # Save the workbook
    wb.save(output_path)
    return output_path

def generate_overview(sheet, data):
    """Generate an overview sheet with key financial metrics."""
    # Set up header
    sheet['A1'] = 'التقرير المالي الشامل | Comprehensive Financial Report'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'المؤشرات المالية الرئيسية | Key Financial Indicators'
    sheet['A3'].font = Font(bold=True, size=14)
    
    # Format cells
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    
    # Add headers
    sheet['A5'] = 'المؤشر | Indicator'
    sheet['B5'] = 'السنة الحالية | Current Year'
    sheet['C5'] = 'السنة السابقة | Previous Year'
    sheet['D5'] = 'التغيير٪ | Change%'
    format_header_row(sheet, row=5)
    
    # Extract key metrics from data
    try:
        metrics = [
            ('إجمالي الإيرادات | Total Revenue', 
             data['income']['إجمالي الإيرادات | Total Revenue']['current'], 
             data['income']['إجمالي الإيرادات | Total Revenue']['previous']),
            ('صافي الربح | Net Profit', 
             data['income']['صافي الربح | Net Profit']['current'], 
             data['income']['صافي الربح | Net Profit']['previous']),
            ('إجمالي الأصول | Total Assets', 
             data['balance']['إجمالي الأصول | Total Assets']['current'], 
             data['balance']['إجمالي الأصول | Total Assets']['previous']),
            ('إجمالي الخصوم | Total Liabilities', 
             data['balance']['إجمالي الخصوم | Total Liabilities']['current'], 
             data['balance']['إجمالي الخصوم | Total Liabilities']['previous']),
            ('إجمالي حقوق الملكية | Total Equity', 
             data['balance']['إجمالي حقوق الملكية | Total Equity']['current'], 
             data['balance']['إجمالي حقوق الملكية | Total Equity']['previous']),
        ]
        
        for i, (metric, current, previous) in enumerate(metrics, start=6):
            sheet[f'A{i}'] = metric
            sheet[f'B{i}'] = current
            sheet[f'C{i}'] = previous
            change_percent = ((current - previous) / previous * 100) if previous else 0
            sheet[f'D{i}'] = f"{change_percent:.2f}%"
            apply_conditional_formatting(sheet, cell=f'D{i}', value=change_percent)
    except Exception as e:
        sheet['A15'] = f"خطأ في حساب المؤشرات: {str(e)}"

def format_header_row(sheet, row):
    """Format a header row with bold font, blue background, and centered alignment."""
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

def apply_conditional_formatting(sheet, cell, value):
    """Apply conditional formatting based on the value."""
    if value > 0:
        sheet[cell].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif value < 0:
        sheet[cell].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def generate_income_statement(sheet, income_data):
    """Generate income statement."""
    sheet['A1'] = 'قائمة الدخل | Income Statement'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'السنة الحالية | Current Year'
    sheet['C3'] = 'السنة السابقة | Previous Year'
    sheet['D3'] = 'التغيير | Change'
    sheet['E3'] = 'التغيير٪ | Change%'
    format_header_row(sheet, row=3)
    
    row = 4
    for item, values in income_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('current', 0)
        sheet[f'C{row}'] = values.get('previous', 0)
        change = values.get('current', 0) - values.get('previous', 0)
        sheet[f'D{row}'] = change
        change_percent = (change / values.get('previous', 1) * 100) if values.get('previous', 0) else 0
        sheet[f'E{row}'] = f"{change_percent:.2f}%"
        if "إجمالي" in item or "صافي" in item:
            for col in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{col}{row}'].font = Font(bold=True)
                sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        row += 1

def generate_balance_sheet(sheet, balance_data):
    """Generate balance sheet."""
    sheet['A1'] = 'قائمة المركز المالي | Balance Sheet'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'السنة الحالية | Current Year'
    sheet['C3'] = 'السنة السابقة | Previous Year'
    sheet['D3'] = 'التغيير | Change'
    sheet['E3'] = 'التغيير٪ | Change%'
    format_header_row(sheet, row=3)
    
    row = 4
    for item, values in balance_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('current', 0)
        sheet[f'C{row}'] = values.get('previous', 0)
        change = values.get('current', 0) - values.get('previous', 0)
        sheet[f'D{row}'] = change
        change_percent = (change / values.get('previous', 1) * 100) if values.get('previous', 0) else 0
        sheet[f'E{row}'] = f"{change_percent:.2f}%"
        if "إجمالي" in item:
            for col in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{col}{row}'].font = Font(bold=True)
                sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        row += 1
    
    # Validate balance sheet (Assets = Liabilities + Equity)
    assets = balance_data.get('إجمالي الأصول | Total Assets', {}).get('current', 0)
    liab_equity = balance_data.get('إجمالي الخصوم وحقوق الملكية | Total Liabilities and Equity', {}).get('current', 0)
    sheet[f'A{row+2}'] = 'التحقق من توازن قائمة المركز المالي | Balance Sheet Check'
    sheet[f'A{row+2}'].font = Font(bold=True)
    if abs(assets - liab_equity) < 0.01:
        sheet[f'B{row+2}'] = 'متوازن ✓ | Balanced ✓'
        sheet[f'B{row+2}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        sheet[f'B{row+2}'] = f'غير متوازن ✗ | Not Balanced ✗ (فرق | Difference: {assets - liab_equity})'
        sheet[f'B{row+2}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def generate_equity_statement(sheet, equity_data):
    """Generate statement of changes in equity."""
    sheet['A1'] = 'قائمة التغيرات في حقوق الملكية | Statement of Changes in Equity'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'رأس المال | Capital'
    sheet['C3'] = 'الاحتياطيات | Reserves'
    sheet['D3'] = 'الأرباح المحتجزة | Retained Earnings'
    sheet['E3'] = 'الإجمالي | Total'
    format_header_row(sheet, row=3)
    
    row = 4
    for item, values in equity_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('capital', 0)
        sheet[f'C{row}'] = values.get('reserves', 0)
        sheet[f'D{row}'] = values.get('retained', 0)
        sheet[f'E{row}'] = values.get('total', 0)
        if "الرصيد في" in item:
            for col in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{col}{row}'].font = Font(bold=True)
                sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        row += 1
    
    # Validate totals
    start_balance = equity_data.get('الرصيد في بداية السنة | Balance at beginning of year', {}).get('total', 0)
    net_profit = equity_data.get('صافي الربح للسنة | Net profit for the year', {}).get('total', 0)
    dividends = equity_data.get('توزيعات الأرباح | Dividends', {}).get('total', 0)
    capital_increase = equity_data.get('زيادة رأس المال | Capital increase', {}).get('total', 0)
    other_changes = equity_data.get('تغييرات أخرى | Other changes', {}).get('total', 0)
    end_balance = equity_data.get('الرصيد في نهاية السنة | Balance at end of year', {}).get('total', 0)
    expected_end = start_balance + net_profit - dividends + capital_increase + other_changes
    sheet[f'A{row+2}'] = 'التحقق من صحة الحسابات | Validation Check'
    sheet[f'A{row+2}'].font = Font(bold=True)
    if abs(expected_end - end_balance) < 0.01:
        sheet[f'B{row+2}'] = 'صحيح ✓ | Correct ✓'
        sheet[f'B{row+2}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        sheet[f'B{row+2}'] = f'غير صحيح ✗ | Incorrect ✗ (فرق | Difference: {expected_end - end_balance})'
        sheet[f'B{row+2}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def generate_cash_flow_statement(sheet, cash_flow_data):
    """Generate cash flow statement."""
    sheet['A1'] = 'قائمة التدفقات النقدية | Cash Flow Statement'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'السنة الحالية | Current Year'
    sheet['C3'] = 'السنة السابقة | Previous Year'
    sheet['D3'] = 'التغيير | Change'
    sheet['E3'] = 'التغيير٪ | Change%'
    format_header_row(sheet, row=3)
    
    row = 4
    for item, values in cash_flow_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('current', 0)
        sheet[f'C{row}'] = values.get('previous', 0)
        change = values.get('current', 0) - values.get('previous', 0)
        sheet[f'D{row}'] = change
        change_percent = (change / values.get('previous', 1) * 100) if values.get('previous', 0) else 0
        sheet[f'E{row}'] = f"{change_percent:.2f}%"
        if "صافي النقد" in item or "النقد وما في حكمه" in item:
            for col in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{col}{row}'].font = Font(bold=True)
                sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        row += 1
    
    # Validate cash flow (cash at beginning + net change = cash at end)
    beg_cash = cash_flow_data.get('النقد وما في حكمه في بداية السنة | Cash and cash equivalents at beginning of year', {}).get('current', 0)
    net_change = cash_flow_data.get('صافي التغير في النقد وما في حكمه | Net change in cash and cash equivalents', {}).get('current', 0)
    end_cash = cash_flow_data.get('النقد وما في حكمه في نهاية السنة | Cash and cash equivalents at end of year', {}).get('current', 0)
    expected_end = beg_cash + net_change
    sheet[f'A{row+2}'] = 'التحقق من صحة حسابات التدفقات النقدية | Cash Flow Validation'
    sheet[f'A{row+2}'].font = Font(bold=True)
    if abs(expected_end - end_cash) < 0.01:
        sheet[f'B{row+2}'] = 'صحيح ✓ | Correct ✓'
        sheet[f'B{row+2}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        sheet[f'B{row+2}'] = f'غير صحيح ✗ | Incorrect ✗ (فرق | Difference: {expected_end - end_cash})'
        sheet[f'B{row+2}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def generate_notes(sheet, notes_data):
    """Generate notes to financial statements."""
    sheet['A1'] = 'الملاحظات على القوائم المالية | Notes to Financial Statements'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 70
    
    notes = [
        (3, 'ملاحظة 1: معلومات عامة | Note 1: General Information', 'note1'),
        (7, 'ملاحظة 2: أسس الإعداد | Note 2: Basis of Preparation', 'note2'),
        (11, 'ملاحظة 3: السياسات المحاسبية الهامة | Note 3: Significant Accounting Policies', 'note3'),
        (15, 'ملاحظة 4: الأحكام والتقديرات المحاسبية الهامة | Note 4: Significant Accounting Judgments and Estimates', 'note4'),
        (19, 'ملاحظة 5: إدارة المخاطر المالية | Note 5: Financial Risk Management', 'note5'),
        (23, 'ملاحظة 6: معلومات إضافية حول بنود القوائم المالية | Note 6: Additional Information on Financial Statement Items', 'note6'),
        (27, 'ملاحظة 7: أحداث لاحقة | Note 7: Subsequent Events', 'note7')
    ]
    for row, title, note_key in notes:
        sheet[f'A{row}'] = title
        sheet[f'A{row}'].font = Font(bold=True)
        if note_key in notes_data and notes_data[note_key]:
            sheet[f'B{row+1}'] = notes_data[note_key]
        else:
            sheet[f'B{row+1}'] = "لم يتم تقديم معلومات. | No information provided."

def generate_charts(sheet, data):
    """Generate financial charts."""
    sheet['A1'] = 'الرسوم البيانية المالية | Financial Charts'
    sheet['A1'].font = Font(bold=True, size=16)
    
    try:
        # Chart 1: Revenue vs Expenses
        revenue_current = data['income'].get('إجمالي الإيرادات | Total Revenue', {}).get('current', 0)
        revenue_previous = data['income'].get('إجمالي الإيرادات | Total Revenue', {}).get('previous', 0)
        expenses_current = data['income'].get('إجمالي المصروفات | Total Expenses', {}).get('current', 0)
        expenses_previous = data['income'].get('إجمالي المصروفات | Total Expenses', {}).get('previous', 0)
        net_profit_current = data['income'].get('صافي الربح | Net Profit', {}).get('current', 0)
        net_profit_previous = data['income'].get('صافي الربح | Net Profit', {}).get('previous', 0)
        
        sheet['A3'] = 'مقارنة الإيرادات والمصروفات | Revenue vs Expenses Comparison'
        sheet['A5'] = 'البند | Item'
        sheet['B5'] = 'السنة الحالية | Current Year'
        sheet['C5'] = 'السنة السابقة | Previous Year'
        sheet['A6'] = 'الإيرادات | Revenue'
        sheet['B6'] = revenue_current
        sheet['C6'] = revenue_previous
        sheet['A7'] = 'المصروفات | Expenses'
        sheet['B7'] = expenses_current
        sheet['C7'] = expenses_previous
        sheet['A8'] = 'صافي الربح | Net Profit'
        sheet['B8'] = net_profit_current
        sheet['C8'] = net_profit_previous
        
        chart1 = BarChart()
        chart1.title = "مقارنة الإيرادات والمصروفات | Revenue vs Expenses"
        chart1.style = 10
        data1 = Reference(sheet, min_col=2, min_row=5, max_row=8, max_col=3)
        cats1 = Reference(sheet, min_col=1, min_row=6, max_row=8)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        sheet.add_chart(chart1, "E3")
        
        # Chart 2: Assets, Liabilities, and Equity
        assets_current = data['balance'].get('إجمالي الأصول | Total Assets', {}).get('current', 0)
        liabilities_current = data['balance'].get('إجمالي الخصوم | Total Liabilities', {}).get('current', 0)
        equity_current = data['balance'].get('إجمالي حقوق الملكية | Total Equity', {}).get('current', 0)
        
        sheet['A12'] = 'مقارنة الأصول والخصوم وحقوق الملكية | Assets, Liabilities and Equity Comparison'
        sheet['A14'] = 'البند | Item'
        sheet['B14'] = 'القيمة | Value'
        sheet['A15'] = 'الأصول | Assets'
        sheet['B15'] = assets_current
        sheet['A16'] = 'الخصوم | Liabilities'
        sheet['B16'] = liabilities_current
        sheet['A17'] = 'حقوق الملكية | Equity'
        sheet['B17'] = equity_current
        
        chart2 = PieChart()
        chart2.title = "توزيع الأصول والخصوم وحقوق الملكية | Distribution of Assets, Liabilities and Equity"
        chart2.style = 10
        data2 = Reference(sheet, min_col=2, min_row=14, max_row=17)
        cats2 = Reference(sheet, min_col=1, min_row=15, max_row=17)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.dLbls = DataLabelList()
        chart2.dLbls.showVal = True
        sheet.add_chart(chart2, "E12")
    except Exception as e:
        sheet['A30'] = f"خطأ في إنشاء الرسوم البيانية: {str(e)}"

# Example usage
if __name__ == "__main__":
    data = {
        'income': {
            'إجمالي الإيرادات | Total Revenue': {'current': 100000, 'previous': 90000},
            'إجمالي المصروفات | Total Expenses': {'current': 80000, 'previous': 75000},
            'صافي الربح | Net Profit': {'current': 20000, 'previous': 15000}
        },
        'balance': {
            'إجمالي الأصول | Total Assets': {'current': 500000, 'previous': 480000},
            'إجمالي الخصوم | Total Liabilities': {'current': 200000, 'previous': 190000},
            'إجمالي حقوق الملكية | Total Equity': {'current': 300000, 'previous': 290000}
        },
        'cash_flow': {
            'النقد وما في حكمه في بداية السنة | Cash and cash equivalents at beginning of year': {'current': 50000, 'previous': 45000},
            'صافي التغير في النقد وما في حكمه | Net change in cash and cash equivalents': {'current': 5000, 'previous': 4000},
            'النقد وما في حكمه في نهاية السنة | Cash and cash equivalents at end of year': {'current': 55000, 'previous': 49000}
        },
        'equity': {
            'الرصيد في بداية السنة | Balance at beginning of year': {'capital': 100000, 'reserves': 50000, 'retained': 150000, 'total': 300000},
            'صافي الربح للسنة | Net profit for the year': {'capital': 0, 'reserves': 0, 'retained': 20000, 'total': 20000},
            'توزيعات الأرباح | Dividends': {'capital': 0, 'reserves': 0, 'retained': -5000, 'total': -5000},
            'زيادة رأس المال | Capital increase': {'capital': 10000, 'reserves': 0, 'retained': 0, 'total': 10000},
            'تغييرات أخرى | Other changes': {'capital': 0, 'reserves': 0, 'retained': 0, 'total': 0},
            'الرصيد في نهاية السنة | Balance at end of year': {'capital': 110000, 'reserves': 50000, 'retained': 165000, 'total': 325000}
        },
        'notes': {
            'note1': 'معلومات عامة عن الشركة.',
            'note2': 'أسس الإعداد تعتمد على المعايير الدولية.',
            'note3': 'السياسات المحاسبية تتبع GAAP.',
            'note4': 'التقديرات تعتمد على بيانات السوق.',
            'note5': 'إدارة المخاطر تشمل التحوط.',
            'note6': 'معلومات إضافية عن البنود.',
            'note7': 'لا يوجد أحداث لاحقة.'
        }
    }
    output_path = "financial_statements_final.xlsx"
    generate_financial_statements(data, output_path)
