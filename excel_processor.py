import os
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart, LineChart, Series

def generate_financial_statements(data, output_path):
    """Generate financial statements based on the provided data."""
    wb = openpyxl.Workbook()
    
    # Create sheets for different financial statements
    sheets = {
        'تقرير عام | Overview': wb.active,
        'قائمة الدخل | Income Statement': wb.create_sheet(),
        'قائمة المركز المالي | Balance Sheet': wb.create_sheet(),
        'قائمة التغيرات في حقوق الملكية | Equity': wb.create_sheet(),
        'قائمة التدفقات النقدية | Cash Flow': wb.create_sheet(),
        'الملاحظات | Notes': wb.create_sheet(),
        'الرسوم البيانية | Charts': wb.create_sheet()
    }
    
    # Rename the default sheet
    sheets['تقرير عام | Overview'].title = 'تقرير عام | Overview'
    
    # Generate each statement
    generate_overview(sheets['تقرير عام | Overview'], data)
    generate_income_statement(sheets['قائمة الدخل | Income Statement'], data['income'])
    generate_balance_sheet(sheets['قائمة المركز المالي | Balance Sheet'], data['balance'])
    generate_equity_statement(sheets['قائمة التغيرات في حقوق الملكية | Equity'], data['equity'])
    generate_cash_flow_statement(sheets['قائمة التدفقات النقدية | Cash Flow'], data['cash_flow'])
    generate_notes(sheets['الملاحظات | Notes'], data['notes'])
    generate_charts(sheets['الرسوم البيانية | Charts'], data)
    
    # Save the workbook
    wb.save(output_path)
    return output_path

def generate_overview(sheet, data):
    """Generate an overview sheet with key financial metrics."""
    # Set up header
    sheet['A1'] = 'التقرير المالي الشامل | Comprehensive Financial Report'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'المؤشرات المالية الرئيسية | Key Financial Indicators'
    sheet['A3'].font = Font(bold=True, size=14)
    
    # Format cells
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    
    # Set up key metrics header
    sheet['A5'] = 'المؤشر | Indicator'
    sheet['B5'] = 'السنة الحالية | Current Year'
    sheet['C5'] = 'السنة السابقة | Previous Year'
    sheet['D5'] = 'التغيير٪ | Change%'
    
    for cell in sheet['5:5']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Extract key metrics from data
    try:
        # Revenue
        total_revenue_current = data['income'].get('إجمالي الإيرادات | Total Revenue', {}).get('current', 0)
        total_revenue_previous = data['income'].get('إجمالي الإيرادات | Total Revenue', {}).get('previous', 0)
        
        # Net profit
        net_profit_current = data['income'].get('صافي الربح | Net Profit', {}).get('current', 0)
        net_profit_previous = data['income'].get('صافي الربح | Net Profit', {}).get('previous', 0)
        
        # Total assets
        total_assets_current = data['balance'].get('إجمالي الأصول | Total Assets', {}).get('current', 0)
        total_assets_previous = data['balance'].get('إجمالي الأصول | Total Assets', {}).get('previous', 0)
        
        # Total liabilities
        total_liabilities_current = data['balance'].get('إجمالي الخصوم | Total Liabilities', {}).get('current', 0)
        total_liabilities_previous = data['balance'].get('إجمالي الخصوم | Total Liabilities', {}).get('previous', 0)
        
        # Total equity
        total_equity_current = data['balance'].get('إجمالي حقوق الملكية | Total Equity', {}).get('current', 0)
        total_equity_previous = data['balance'].get('إجمالي حقوق الملكية | Total Equity', {}).get('previous', 0)
        
        # Cash at end of year
        cash_end_current = data['cash_flow'].get('النقد وما في حكمه في نهاية السنة | Cash and cash equivalents at end of year', {}).get('current', 0)
        cash_end_previous = data['cash_flow'].get('النقد وما في حكمه في نهاية السنة | Cash and cash equivalents at end of year', {}).get('previous', 0)
        
        # Calculate ratios
        profitability_current = (net_profit_current / total_revenue_current * 100) if total_revenue_current else 0
        profitability_previous = (net_profit_previous / total_revenue_previous * 100) if total_revenue_previous else 0
        
        liquidity_current = total_assets_current / total_liabilities_current if total_liabilities_current else 0
        liquidity_previous = total_assets_previous / total_liabilities_previous if total_liabilities_previous else 0
        
        debt_equity_current = total_liabilities_current / total_equity_current if total_equity_current else 0
        debt_equity_previous = total_liabilities_previous / total_equity_previous if total_equity_previous else 0
        
        # Calculate percentage changes
        def calculate_change(current, previous):
            if previous:
                return ((current - previous) / previous) * 100
            return 0
        
        revenue_change = calculate_change(total_revenue_current, total_revenue_previous)
        profit_change = calculate_change(net_profit_current, net_profit_previous)
        assets_change = calculate_change(total_assets_current, total_assets_previous)
        liabilities_change = calculate_change(total_liabilities_current, total_liabilities_previous)
        equity_change = calculate_change(total_equity_current, total_equity_previous)
        cash_change = calculate_change(cash_end_current, cash_end_previous)
        profitability_change = calculate_change(profitability_current, profitability_previous)
        liquidity_change = calculate_change(liquidity_current, liquidity_previous)
        debt_equity_change = calculate_change(debt_equity_current, debt_equity_previous)
        
        # Add metrics to sheet
        metrics = [
            ('إجمالي الإيرادات | Total Revenue', total_revenue_current, total_revenue_previous, revenue_change),
            ('صافي الربح | Net Profit', net_profit_current, net_profit_previous, profit_change),
            ('إجمالي الأصول | Total Assets', total_assets_current, total_assets_previous, assets_change),
            ('إجمالي الخصوم | Total Liabilities', total_liabilities_current, total_liabilities_previous, liabilities_change),
            ('إجمالي حقوق الملكية | Total Equity', total_equity_current, total_equity_previous, equity_change),
            ('النقد في نهاية السنة | Cash at End of Year', cash_end_current, cash_end_previous, cash_change),
            ('معدل الربحية٪ | Profitability Ratio %', profitability_current, profitability_previous, profitability_change),
            ('نسبة السيولة | Liquidity Ratio', liquidity_current, liquidity_previous, liquidity_change),
            ('نسبة الدين إلى حقوق الملكية | Debt to Equity', debt_equity_current, debt_equity_previous, debt_equity_change)
        ]
        
        for i, (metric, current, previous, change) in enumerate(metrics, start=6):
            sheet[f'A{i}'] = metric
            sheet[f'B{i}'] = current
            sheet[f'C{i}'] = previous
            sheet[f'D{i}'] = f"{change:.2f}%"
            
            # Color code changes
            if change > 0 and i < 9:  # For ratios, the meaning of positive/negative can be different
                sheet[f'D{i}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif change < 0 and i < 9:
                sheet[f'D{i}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    except Exception as e:
        sheet['A15'] = f"خطأ في حساب المؤشرات: {str(e)}"
    
    # Add a financial summary section
    sheet['A16'] = 'ملخص الأداء المالي | Financial Performance Summary'
    sheet['A16'].font = Font(bold=True, size=14)
    
    try:
        if net_profit_current > net_profit_previous:
            performance = "تحسن الأداء المالي مقارنة بالعام السابق. | Financial performance improved compared to previous year."
        elif net_profit_current < net_profit_previous:
            performance = "انخفاض الأداء المالي مقارنة بالعام السابق. | Financial performance declined compared to previous year."
        else:
            performance = "استقرار الأداء المالي مقارنة بالعام السابق. | Financial performance stable compared to previous year."
        
        sheet['A18'] = performance
        
        # Add liquidity assessment
        if liquidity_current >= 2:
            liquidity_assessment = "وضع السيولة ممتاز. | Excellent liquidity position."
        elif liquidity_current >= 1:
            liquidity_assessment = "وضع السيولة جيد. | Good liquidity position."
        else:
            liquidity_assessment = "وضع السيولة يحتاج إلى تحسين. | Liquidity position needs improvement."
        
        sheet['A19'] = liquidity_assessment
        
        # Add debt assessment
        if debt_equity_current <= 0.5:
            debt_assessment = "نسبة الدين منخفضة، مما يشير إلى مخاطر مالية منخفضة. | Low debt ratio indicating low financial risk."
        elif debt_equity_current <= 1:
            debt_assessment = "نسبة الدين معتدلة. | Moderate debt ratio."
        else:
            debt_assessment = "نسبة الدين مرتفعة، مما قد يشير إلى مخاطر مالية. | High debt ratio which may indicate financial risk."
        
        sheet['A20'] = debt_assessment
    
    except Exception as e:
        sheet['A18'] = f"خطأ في تحليل الأداء: {str(e)}"

def generate_income_statement(sheet, income_data):
    """Generate income statement."""
    # Set up header
    sheet['A1'] = 'قائمة الدخل | Income Statement'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'السنة الحالية | Current Year'
    sheet['C3'] = 'السنة السابقة | Previous Year'
    sheet['D3'] = 'التغيير | Change'
    sheet['E3'] = 'التغيير٪ | Change%'
    
    # Format header row
    for cell in sheet['3:3']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    
    # Add income items
    row = 4
    for item, values in income_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('current', 0)
        sheet[f'C{row}'] = values.get('previous', 0)
        
        # Calculate change
        current = values.get('current', 0)
        previous = values.get('previous', 0)
        change = current - previous
        sheet[f'D{row}'] = change
        
        # Calculate percentage change
        if previous != 0:
            change_percent = (change / previous) * 100
            sheet[f'E{row}'] = f"{change_percent:.2f}%"
        else:
            sheet[f'E{row}'] = "N/A"
        
        # Format totals and net profit
        if "إجمالي" in item or "صافي" in item or "الربح" in item:
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'].font = Font(bold=True)
            sheet[f'C{row}'].font = Font(bold=True)
            sheet[f'D{row}'].font = Font(bold=True)
            sheet[f'E{row}'].font = Font(bold=True)
            
            # Add background color
            for col in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        row += 1

def generate_balance_sheet(sheet, balance_data):
    """Generate balance sheet."""
    # Set up header
    sheet['A1'] = 'قائمة المركز المالي | Balance Sheet'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'السنة الحالية | Current Year'
    sheet['C3'] = 'السنة السابقة | Previous Year'
    sheet['D3'] = 'التغيير | Change'
    sheet['E3'] = 'التغيير٪ | Change%'
    
    # Format header row
    for cell in sheet['3:3']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    
    # Add balance sheet items
    row = 4
    for item, values in balance_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('current', 0)
        sheet[f'C{row}'] = values.get('previous', 0)
        
        # Calculate change
        current = values.get('current', 0)
        previous = values.get('previous', 0)
        change = current - previous
        sheet[f'D{row}'] = change
        
        # Calculate percentage change
        if previous != 0:
            change_percent = (change / previous) * 100
            sheet[f'E{row}'] = f"{change_percent:.2f}%"
        else:
            sheet[f'E{row}'] = "N/A"
        
        # Format section headers and totals
        if "إجمالي" in item or "الأصول" in item or "الخصوم وحقوق الملكية" in item or "الخصوم المتداولة" in item or "الخصوم غير المتداولة" in item or "حقوق الملكية" in item:
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'].font = Font(bold=True)
            sheet[f'C{row}'].font = Font(bold=True)
            sheet[f'D{row}'].font = Font(bold=True)
            sheet[f'E{row}'].font = Font(bold=True)
            
            # Add background color for totals
            if "إجمالي" in item:
                for col in ['A', 'B', 'C', 'D', 'E']:
                    sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        row += 1
    
    # Validate balance sheet (Assets = Liabilities + Equity)
    try:
        assets = balance_data.get('إجمالي الأصول | Total Assets', {}).get('current', 0)
        liab_equity = balance_data.get('إجمالي الخصوم وحقوق الملكية | Total Liabilities and Equity', {}).get('current', 0)
        
        sheet[f'A{row+2}'] = 'التحقق من توازن قائمة المركز المالي | Balance Sheet Check'
        sheet[f'A{row+2}'].font = Font(bold=True)
        
        if abs(assets - liab_equity) < 0.01:  # Allow for floating point imprecision
            sheet[f'B{row+2}'] = 'متوازن ✓ | Balanced ✓'
            sheet[f'B{row+2}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            sheet[f'B{row+2}'] = f'غير متوازن ✗ | Not Balanced ✗ (فرق | Difference: {assets - liab_equity})'
            sheet[f'B{row+2}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    except:
        pass

def generate_equity_statement(sheet, equity_data):
    """Generate statement of changes in equity."""
    # Set up header
    sheet['A1'] = 'قائمة التغيرات في حقوق الملكية | Statement of Changes in Equity'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'رأس المال | Capital'
    sheet['C3'] = 'الاحتياطيات | Reserves'
    sheet['D3'] = 'الأرباح المحتجزة | Retained Earnings'
    sheet['E3'] = 'الإجمالي | Total'
    
    # Format header row
    for cell in sheet['3:3']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    
    # Add equity items
    row = 4
    for item, values in equity_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('capital', 0)
        sheet[f'C{row}'] = values.get('reserves', 0)
        sheet[f'D{row}'] = values.get('retained', 0)
        sheet[f'E{row}'] = values.get('total', 0)
        
        # Format beginning and ending balances
        if "الرصيد في" in item:
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'].font = Font(bold=True)
            sheet[f'C{row}'].font = Font(bold=True)
            sheet[f'D{row}'].font = Font(bold=True)
            sheet[f'E{row}'].font = Font(bold=True)
            
            # Add background color
            for col in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        row += 1
    
    # Validate totals
    try:
        start_balance = equity_data.get('الرصيد في بداية السنة | Balance at beginning of year', {}).get('total', 0)
        net_profit = equity_data.get('صافي الربح للسنة | Net profit for the year', {}).get('total', 0)
        dividends = equity_data.get('توزيعات الأرباح | Dividends', {}).get('total', 0)
        capital_increase = equity_data.get('زيادة رأس المال | Capital increase', {}).get('total', 0)
        other_changes = equity_data.get('تغييرات أخرى | Other changes', {}).get('total', 0)
        end_balance = equity_data.get('الرصيد في نهاية السنة | Balance at end of year', {}).get('total', 0)
        
        expected_end = start_balance + net_profit - dividends + capital_increase + other_changes
        
        sheet[f'A{row+2}'] = 'التحقق من صحة الحسابات | Validation Check'
        sheet[f'A{row+2}'].font = Font(bold=True)
        
        if abs(expected_end - end_balance) < 0.01:  # Allow for floating point imprecision
            sheet[f'B{row+2}'] = 'صحيح ✓ | Correct ✓'
            sheet[f'B{row+2}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            sheet[f'B{row+2}'] = f'غير صحيح ✗ | Incorrect ✗ (فرق | Difference: {expected_end - end_balance})'
            sheet[f'B{row+2}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    except:
        pass

def generate_cash_flow_statement(sheet, cash_flow_data):
    """Generate cash flow statement."""
    # Set up header
    sheet['A1'] = 'قائمة التدفقات النقدية | Cash Flow Statement'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A3'] = 'البند | Item'
    sheet['B3'] = 'السنة الحالية | Current Year'
    sheet['C3'] = 'السنة السابقة | Previous Year'
    sheet['D3'] = 'التغيير | Change'
    sheet['E3'] = 'التغيير٪ | Change%'
    
    # Format header row
    for cell in sheet['3:3']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    
    # Add cash flow items
    row = 4
    for item, values in cash_flow_data.items():
        sheet[f'A{row}'] = item
        sheet[f'B{row}'] = values.get('current', 0)
        sheet[f'C{row}'] = values.get('previous', 0)
        
        # Calculate change
        current = values.get('current', 0)
        previous = values.get('previous', 0)
        change = current - previous
        sheet[f'D{row}'] = change
        
        # Calculate percentage change
        if previous != 0:
            change_percent = (change / previous) * 100
            sheet[f'E{row}'] = f"{change_percent:.2f}%"
        else:
            sheet[f'E{row}'] = "N/A"
        
        # Format section headers and net cash
        if "التدفقات النقدية من" in item or "صافي النقد" in item or "النقد وما في حكمه" in item:
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'].font = Font(bold=True)
            sheet[f'C{row}'].font = Font(bold=True)
            sheet[f'D{row}'].font = Font(bold=True)
            sheet[f'E{row}'].font = Font(bold=True)
            
            # Add background color for net cash and cash at year-end
            if "صافي النقد" in item or "النقد وما في حكمه في نهاية السنة" in item:
                for col in ['A', 'B', 'C', 'D', 'E']:
                    sheet[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        row += 1
    
    # Validate cash flow (cash at beginning + net change = cash at end)
    try:
        beg_cash = cash_flow_data.get('النقد وما في حكمه في بداية السنة | Cash and cash equivalents at beginning of year', {}).get('current', 0)
        net_change = cash_flow_data.get('صافي التغير في النقد وما في حكمه | Net change in cash and cash equivalents', {}).get('current', 0)
        end_cash = cash_flow_data.get('النقد وما في حكمه في نهاية السنة | Cash and cash equivalents at end of year', {}).get('current', 0)
        
        expected_end = beg_cash + net_change
        
        sheet[f'A{row+2}'] = 'التحقق من صحة حسابات التدفقات النقدية | Cash Flow Validation'
        sheet[f'A{row+2}'].font = Font(bold=True)
        
        if abs(expected_end - end_cash) < 0.01:  # Allow for floating point imprecision
            sheet[f'B{row+2}'] = 'صحيح ✓ | Correct ✓'
            sheet[f'B{row+2}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            sheet[f'B{row+2}'] = f'غير صحيح ✗ | Incorrect ✗ (فرق | Difference: {expected_end - end_cash})'
            sheet[f'B{row+2}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    except:
        pass

def generate_notes(sheet, notes_data):
    """Generate notes to financial statements."""
    # Set up header
    sheet['A1'] = 'الملاحظات على القوائم المالية | Notes to Financial Statements'
    sheet['A1'].font = Font(bold=True, size=16)
    
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 70
    
    # Add notes
    notes = [
        (3, 'ملاحظة 1: معلومات عامة | Note 1: General Information', 'note1'),
        (7, 'ملاحظة 2: أسس الإعداد | Note 2: Basis of Preparation', 'note2'),
        (11, 'ملاحظة 3: السياسات المحاسبية الهامة | Note 3: Significant Accounting Policies', 'note3'),
        (15, 'ملاحظة 4: الأحكام والتقديرات المحاسبية الهامة | Note 4: Significant Accounting Judgments and Estimates', 'note4'),
        (19, 'ملاحظة 5: إدارة المخاطر المالية | Note 5: Financial Risk Management', 'note5'),
        (23, 'ملاحظة 6: معلومات إضافية حول بنود القوائم المالية | Note 6: Additional Information on Financial Statement Items', 'note6'),
        (27, 'ملاحظة 7: أحداث لاحقة | Note 7: Subsequent Events', 'note7')
    ]
    
    for row, title, note_key in notes:
        sheet[f'A{row}'] = title
        sheet[f'A{row}'].font = Font(bold=True)
        
        # Add note content
        if note_key in notes_data and notes_data[note_key]:
            sheet[f'B{row+1}'] = notes_data[note_key]
        else:
            sheet[f'B{row+1}'] = "لم يتم تقديم معلومات. | No information provided."

def generate_charts(sheet, data):
    """Generate financial charts."""
    # Set up header
    sheet['A1'] = 'الرسوم البيانية المالية | Financial Charts'
    sheet['A1'].font = Font(bold=True, size=16)
    
    try:
        # Extract data for charts
        income_data = data['income']
        balance_data = data['balance']
        cash_flow_data = data['cash_flow']
        
        # Create revenue vs expenses chart
        revenue_current = income_data.get('إجمالي الإيرادات | Total Revenue', {}).get('current', 0)
        revenue_previous = income_data.get('إجمالي الإيرادات | Total Revenue', {}).get('previous', 0)
        
        expenses_current = income_data.get('إجمالي المصروفات | Total Expenses', {}).get('current', 0)
        expenses_previous = income_data.get('إجمالي المصروفات | Total Expenses', {}).get('previous', 0)
        
        net_profit_current = income_data.get('صافي الربح | Net Profit', {}).get('current', 0)
        net_profit_previous = income_data.get('صافي الربح | Net Profit', {}).get('previous', 0)
        
        # Add data for chart 1
        sheet['A3'] = 'مقارنة الإيرادات والمصروفات | Revenue vs Expenses Comparison'
        sheet['A3'].font = Font(bold=True)
        
        sheet['A5'] = 'البند | Item'
        sheet['B5'] = 'السنة الحالية | Current Year'
        sheet['C5'] = 'السنة السابقة | Previous Year'
        
        sheet['A6'] = 'الإيرادات | Revenue'
        sheet['B6'] = revenue_current
        sheet['C6'] = revenue_previous
        
        sheet['A7'] = 'المصروفات | Expenses'
        sheet['B7'] = expenses_current
        sheet['C7'] = expenses_previous
        
        sheet['A8'] = 'صافي الربح | Net Profit'
        sheet['B8'] = net_profit_current
        sheet['C8'] = net_profit_previous
        
        # Create chart 1
        chart1 = BarChart()
        chart1.title = "مقارنة الإيرادات والمصروفات | Revenue vs Expenses"
        chart1.style = 10
        chart1.x_axis.title = "البند | Item"
        chart1.y_axis.title = "القيمة | Value"
        
        data1 = Reference(sheet, min_col=2, min_row=5, max_row=8, max_col=3)
        cats1 = Reference(sheet, min_col=1, min_row=6, max_row=8)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.shape = 4
        sheet.add_chart(chart1, "E3")
        
        # Add data for chart 2 - Assets, Liabilities and Equity
        sheet['A12'] = 'مقارنة الأصول والخصوم وحقوق الملكية | Assets, Liabilities and Equity Comparison'
        sheet['A12'].font = Font(bold=True)
        
        assets_current = balance_data.get('إجمالي الأصول | Total Assets', {}).get('current', 0)
        liabilities_current = balance_data.get('إجمالي الخصوم | Total Liabilities', {}).get('current', 0)
        equity_current = balance_data.get('إجمالي حقوق الملكية | Total Equity', {}).get('current', 0)
        
        sheet['A14'] = 'البند | Item'
        sheet['B14'] = 'القيمة | Value'
        
        sheet['A15'] = 'الأصول | Assets'
        sheet['B15'] = assets_current
        
        sheet['A16'] = 'الخصوم | Liabilities'
        sheet['B16'] = liabilities_current
        
        sheet['A17'] = 'حقوق الملكية | Equity'
        sheet['B17'] = equity_current
        
        # Create chart 2
        chart2 = PieChart()
        chart2.title = "توزيع الأصول والخصوم وحقوق الملكية | Distribution of Assets, Liabilities and Equity"
        chart2.style = 10
        
        data2 = Reference(sheet, min_col=2, min_row=14, max_row=17)
        cats2 = Reference(sheet, min_col=1, min_row=15, max_row=17)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.dataLabels = True
        sheet.add_chart(chart2, "E12")
        
        # Add data for chart 3 - Cash Flow Comparison
        sheet['A21'] = 'مقارنة التدفقات النقدية | Cash Flow Comparison'
        sheet['A21'].font = Font(bold=True)
        
        operating_current = cash_flow_data.get('صافي النقد من الأنشطة التشغيلية | Net cash from operating activities', {}).get('current', 0)
        investing_current = cash_flow_data.get('صافي النقد من الأنشطة الاستثمارية | Net cash from investing activities', {}).get('current', 0)
        financing_current = cash_flow_data.get('صافي النقد من الأنشطة التمويلية | Net cash from financing activities', {}).get('current', 0)
        
        operating_previous = cash_flow_data.get('صافي النقد من الأنشطة التشغيلية | Net cash from operating activities', {}).get('previous', 0)
        investing_previous = cash_flow_data.get('صافي النقد من الأنشطة الاستثمارية | Net cash from investing activities', {}).get('previous', 0)
        financing_previous = cash_flow_data.get('صافي النقد من الأنشطة التمويلية | Net cash from financing activities', {}).get('previous', 0)
        
        sheet['A23'] = 'مصدر التدفق النقدي | Cash Flow Source'
        sheet['B23'] = 'السنة الحالية | Current Year'
        sheet['C23'] = 'السنة السابقة | Previous Year'
        
        sheet['A24'] = 'الأنشطة التشغيلية | Operating Activities'
        sheet['B24'] = operating_current
        sheet['C24'] = operating_previous
        
        sheet['A25'] = 'الأنشطة الاستثمارية | Investing Activities'
        sheet['B25'] = investing_current
        sheet['C25'] = investing_previous
        
        sheet['A26'] = 'الأنشطة التمويلية | Financing Activities'
        sheet['B26'] = financing_current
        sheet['C26'] = financing_previous
        
        # Create chart 3
        chart3 = BarChart()
        chart3.title = "مقارنة التدفقات النقدية | Cash Flow Comparison"
        chart3.style = 10
        chart3.x_axis.title = "مصدر التدفق النقدي | Cash Flow Source"
        chart3.y_axis.title = "القيمة | Value"
        
        data3 = Reference(sheet, min_col=2, min_row=23, max_row=26, max_col=3)
        cats3 = Reference(sheet, min_col=1, min_row=24, max_row=26)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        sheet.add_chart(chart3, "E21")
        
    except Exception as e:
        sheet['A30'] = f"خطأ في إنشاء الرسوم البيانية: {str(e)}"
